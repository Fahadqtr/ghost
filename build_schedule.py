# -*- coding: utf-8 -*-
"""
بناء ملف إكسل لجدول الورديات وحجز الإجازات
قسم العمليات الجمركية - فريق الوردية
نظام: 6 أيام عمل + 4 أيام راحة (دورية) — بداية الدوام 2:00 عصراً
مع نظام حجز إجازات يكشف التعارض تلقائياً.
"""
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.hyperlink import Hyperlink

# ------------------------------------------------------------------ بيانات
EMPLOYEES = [
    (1, "سالم شليويح المري", 392),
    (2, "فهد عبدالعزيز عبدالله", 2306),
    (3, "محمد ناصر الغانم", 4756),
    (4, "راشد عيسى التميمي", 4749),
    (5, "طلال زايد المري", 2927),
    (6, "منصور مبارك الجابري", 4522),
    (7, "محمد حمد الجابر", 3317),
    (8, "عبدالله سعد المهندي", 2424),
]

LEAVE_TYPES = ["سنوية", "عارض", "دورية", "مرضية", "مرافق مريض", "أخرى"]
STATUSES = ["معتمد", "قيد الانتظار", "مرفوض"]

# بيانات الإجازات المسجّلة مسبقاً (مأخوذة من ورقة المستخدم) — (الاسم، النوع، من، إلى، الحالة)
import datetime as _dt
LEAVE_DATA = [
    ("سالم شليويح المري", "دورية", _dt.date(2026, 6, 17), _dt.date(2026, 6, 19), "معتمد"),
    ("فهد عبدالعزيز عبدالله", "سنوية", _dt.date(2026, 6, 18), _dt.date(2026, 6, 19), "معتمد"),
    ("محمد ناصر الغانم", "دورية", _dt.date(2026, 6, 28), _dt.date(2026, 7, 2), "معتمد"),
    ("محمد ناصر الغانم", "دورية", _dt.date(2026, 7, 5), _dt.date(2026, 7, 9), "معتمد"),
    ("منصور مبارك الجابري", "سنوية", _dt.date(2026, 7, 4), _dt.date(2026, 7, 9), "معتمد"),
    ("منصور مبارك الجابري", "سنوية", _dt.date(2026, 6, 18), _dt.date(2026, 6, 19), "معتمد"),
    ("محمد حمد الجابر", "سنوية", _dt.date(2026, 6, 17), _dt.date(2026, 6, 19), "معتمد"),
    ("عبدالله سعد المهندي", "مرافق مريض", _dt.date(2026, 6, 17), _dt.date(2026, 7, 30), "معتمد"),
    ("راشد عيسى التميمي", "سنوية", _dt.date(2026, 7, 4), _dt.date(2026, 7, 9), "معتمد"),
    ("راشد عيسى التميمي", "عارض", _dt.date(2026, 7, 14), _dt.date(2026, 7, 14), "معتمد"),
    ("راشد عيسى التميمي", "سنوية", _dt.date(2026, 7, 15), _dt.date(2026, 7, 19), "معتمد"),
    ("راشد عيسى التميمي", "أخرى", _dt.date(2026, 6, 17), _dt.date(2026, 6, 17), "معتمد"),
]

# الورديات (الزامات) وأوقاتها
SHIFTS = [
    ("صباح", "6:00 ص ← 1:00 م"),
    ("عصر", "1:00 م ← 9:00 م"),
    ("ليل", "9:00 م ← 6:00 ص (اليوم التالي)"),
]
SHIFT_NAMES = [s[0] for s in SHIFTS]
DEFAULT_SHIFT = "عصر"
# ترتيب تدوير الورديات: كل دورة عمل (6 أيام) تنتقل للوردية التالية
ROTATION = ["صباح", "عصر", "ليل"]

# لون لكل وردية / حالة / نوع إجازة
COLORS = {
    "صباح":       "FFF2CC",   # أصفر باهت
    "عصر":        "C6EFCE",   # أخضر فاتح
    "ليل":        "8EAADB",   # أزرق
    "راحة":       "D9D9D9",   # رمادي
    "سنوية":      "FFE699",   # كهرماني
    "عارض":       "F8CBAD",   # برتقالي
    "دورية":      "BDD7EE",   # أزرق فاتح
    "مرضية":      "FFC7CE",   # أحمر فاتح
    "مرافق مريض": "E2A9F3",   # بنفسجي
    "أخرى":       "FCE4D6",   # خوخي
    "عطلة":       "F4B084",   # برتقالي داكن (عطلة رسمية)
}

# العطل الرسمية الافتراضية (قابلة للتعديل في ورقة الإعدادات) — (التاريخ، المناسبة)
HOLIDAYS = [
    (datetime.date(2026, 12, 18), "اليوم الوطني"),
    (datetime.date(2027, 2, 9), "اليوم الرياضي للدولة"),
]

PERIOD_START = datetime.date(2026, 6, 14)   # أول يوم في الجدول (أول وردية صباح)
DAYS = 365                                  # سنة كاملة
AR_DAYS = ["الإثنين", "الثلاثاء", "الأربعاء", "الخميس", "الجمعة", "السبت", "الأحد"]
AR_MONTHS = ["يناير", "فبراير", "مارس", "أبريل", "مايو", "يونيو",
             "يوليو", "أغسطس", "سبتمبر", "أكتوبر", "نوفمبر", "ديسمبر"]

L_FIRST, L_ROWS = 5, 100        # صفوف بيانات حجز الإجازات
L_LAST = L_FIRST + L_ROWS - 1

# ------------------------------------------------------------------ أنماط
THIN = Side(style="thin", color="9BB0C4")
MED = Side(style="medium", color="44546A")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_M = Border(left=MED, right=MED, top=MED, bottom=MED)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)
HDR_FILL = PatternFill("solid", fgColor="44546A")
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Arial", bold=True, color="1F3864", size=16)
SUB_FONT = Font(name="Arial", bold=True, color="44546A", size=12)
BASE_FONT = Font(name="Arial", size=11)


def style_header(cell):
    cell.fill = HDR_FILL
    cell.font = HDR_FONT
    cell.alignment = CENTER
    cell.border = BORDER


def cf_fill(color):
    """تعبئة للتنسيق الشرطي: إكسل يقرأ لون التعبئة من bgColor وليس fgColor."""
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


wb = Workbook()

# =================================================================  تعليمات
ws_info = wb.active
ws_info.title = "تعليمات"
ws_info.sheet_view.rightToLeft = True
ws_info.column_dimensions["A"].width = 3
ws_info.column_dimensions["B"].width = 95

rows = [
    ("title", "نظام جدول الورديات وحجز الإجازات"),
    ("sub", "قسم العمليات الجمركية — فريق الوردية"),
    ("", ""),
    ("h", "محتويات الملف (الأوراق بالأسفل):"),
    ("b", "1) الموظفون: قائمة أفراد الفريق وأرقامهم الوظيفية وبداية دورة العمل لكل فرد."),
    ("b", "2) جدول الورديات: تقويم يومي يحسب تلقائياً (عمل / راحة) ويُظهر الإجازات المعتمدة بالألوان."),
    ("b", "3) تقويم الإجازات: عرض واضح للإجازات المعتمدة فقط (موظفون × أيام) مع إجمالي أيام كل موظف."),
    ("b", "4) حجز الإجازات: تُسجَّل فيه طلبات الإجازة، ويكشف التعارض تلقائياً."),
    ("b", "5) الإعدادات والقوائم: ضبط بداية الدورة والحد الأدنى للتغطية وأنواع الإجازات."),
    ("", ""),
    ("h", "نظام الدوام:"),
    ("b", "• 6 أيام عمل متتالية ثم 4 أيام راحة (دورية) — دورة كل 10 أيام."),
    ("b", "• ثلاث ورديات (زامات): صباح ← عصر ← ليل."),
    ("b", "     صباح: 6:00 ص ← 1:00 م   |   عصر: 1:00 م ← 9:00 م   |   ليل: 9:00 م ← 6:00 ص (اليوم التالي)."),
    ("b", "• داخل كل دورة عمل (6 أيام): يومان صباح + يومان عصر + يومان ليل، ثم 4 أيام راحة — ويتكرر النمط."),
    ("b", "• «وردية بداية الفريق» في ورقة الإعدادات تحدّد أول وردية في الدورة — غيّرها لإزاحة الترتيب."),
    ("", ""),
    ("h", "أنواع الإجازات:"),
    ("b", "سنوية • عارض • دورية • مرضية • مرافق مريض • أخرى"),
    ("", ""),
    ("h", "كيف أحجز إجازة بدون تعارض؟"),
    ("b", "1) افتح ورقة «حجز الإجازات»."),
    ("b", "2) اختر اسم الموظف ونوع الإجازة من القائمة المنسدلة، وأدخل تاريخ البداية والنهاية."),
    ("b", "3) عمود «عدد الأيام» يُحسب تلقائياً، وعمود «فحص التعارض» يتحقق فوراً:"),
    ("b", "     ✓ سليم = لا يوجد تعارض   |   ⚠ تعارض = للموظف إجازة أخرى متداخلة بنفس الفترة."),
    ("b", "4) عمود «تنبيه التغطية» ينبّه إذا تجاوز عدد المُجازين في نفس اليوم الحد المسموح."),
    ("b", "5) تظهر الإجازة ملوّنة في جدول الورديات والتقويم فور إدخالها (إلا إذا كانت حالتها «مرفوض»)."),
    ("", ""),
    ("h", "ملاحظات:"),
    ("b", "• الصفوف الحمراء/التنبيهات تعني وجود تعارض — عالجه قبل الاعتماد."),
    ("b", "• في «جدول الورديات» يظهر أسفل كل يوم: عدد العاملين وعدد المُجازين؛ ويتلوّن بالأحمر عند نقص التغطية."),
    ("b", "• كل التواريخ بصيغة يوم/شهر/سنة. عدّل «بداية الفترة» وعدد الموظفين بحرية."),
]
r = 1
for kind, text in rows:
    c = ws_info.cell(row=r, column=2, value=text)
    if kind == "title":
        c.font = TITLE_FONT
        ws_info.row_dimensions[r].height = 26
    elif kind == "sub":
        c.font = SUB_FONT
    elif kind == "h":
        c.font = Font(name="Arial", bold=True, color="C00000", size=12)
    else:
        c.font = BASE_FONT
    c.alignment = RIGHT
    r += 1

# =================================================================  الإعدادات والقوائم
ws_set = wb.create_sheet("الإعدادات والقوائم")
ws_set.sheet_view.rightToLeft = True
ws_set.column_dimensions["A"].width = 30
ws_set.column_dimensions["B"].width = 18
ws_set.column_dimensions["D"].width = 18
ws_set.column_dimensions["F"].width = 18

ws_set["A1"] = "الإعدادات العامة"
ws_set["A1"].font = SUB_FONT
settings = [
    ("بداية فترة الجدول", PERIOD_START),
    ("طول دورة العمل (أيام)", 6),
    ("طول الراحة الدورية (أيام)", 4),
    ("الحد الأدنى للعاملين يومياً", 4),
    ("أقصى عدد مُجازين في اليوم", 3),
    ("وردية بداية الفريق", "صباح"),
]
for i, (k, v) in enumerate(settings, start=2):
    a = ws_set.cell(row=i, column=1, value=k)
    a.font = Font(name="Arial", bold=True, size=11)
    a.alignment = RIGHT
    a.border = BORDER
    b = ws_set.cell(row=i, column=2, value=v)
    b.alignment = CENTER
    b.border = BORDER
    if isinstance(v, datetime.date):
        b.number_format = "DD/MM/YYYY"
# مراجع الإعدادات — أسماء معرّفة لضمان التحديث في الصيغ والتنسيق الشرطي عبر الأوراق
from openpyxl.workbook.defined_name import DefinedName
wb.defined_names["CycleWork"] = DefinedName("CycleWork", attr_text="'الإعدادات والقوائم'!$B$3")
wb.defined_names["CycleRest"] = DefinedName("CycleRest", attr_text="'الإعدادات والقوائم'!$B$4")
wb.defined_names["MinStaff"] = DefinedName("MinStaff", attr_text="'الإعدادات والقوائم'!$B$5")
wb.defined_names["MaxLeave"] = DefinedName("MaxLeave", attr_text="'الإعدادات والقوائم'!$B$6")
wb.defined_names["StartShift"] = DefinedName("StartShift", attr_text="'الإعدادات والقوائم'!$B$7")
CYCLE_WORK = "CycleWork"
CYCLE_REST = "CycleRest"
MIN_STAFF = "MinStaff"
MAX_LEAVE = "MaxLeave"
START_SHIFT = "StartShift"

# قوائم منسدلة
ws_set["D1"] = "أنواع الإجازات"
ws_set["D1"].font = SUB_FONT
for i, t in enumerate(LEAVE_TYPES, start=2):
    cc = ws_set.cell(row=i, column=4, value=t)
    cc.alignment = CENTER
    cc.border = BORDER
    cc.fill = PatternFill("solid", fgColor=COLORS[t])

ws_set["F1"] = "حالات الطلب"
ws_set["F1"].font = SUB_FONT
for i, s in enumerate(STATUSES, start=2):
    cc = ws_set.cell(row=i, column=6, value=s)
    cc.alignment = CENTER
    cc.border = BORDER

TYPES_RANGE = f"'الإعدادات والقوائم'!$D$2:$D${1+len(LEAVE_TYPES)}"
STATUS_RANGE = f"'الإعدادات والقوائم'!$F$2:$F${1+len(STATUSES)}"
# أسماء معرّفة للقوائم لضمان عمل القوائم المنسدلة في كل إصدارات إكسل
wb.defined_names["LeaveTypes"] = DefinedName("LeaveTypes", attr_text=TYPES_RANGE)
wb.defined_names["ReqStatus"] = DefinedName("ReqStatus", attr_text=STATUS_RANGE)
TYPES_RANGE = "LeaveTypes"
STATUS_RANGE = "ReqStatus"

# جدول الورديات (الزامات) وأوقاتها
ws_set["H1"] = "الوردية"
ws_set["I1"] = "التوقيت"
ws_set.column_dimensions["H"].width = 12
ws_set.column_dimensions["I"].width = 30
for cc in (ws_set["H1"], ws_set["I1"]):
    style_header(cc)
for i, (nm, tm) in enumerate(SHIFTS, start=2):
    h = ws_set.cell(row=i, column=8, value=nm)
    h.alignment = CENTER
    h.border = BORDER
    h.fill = PatternFill("solid", fgColor=COLORS[nm])
    t = ws_set.cell(row=i, column=9, value=tm)
    t.alignment = RIGHT
    t.border = BORDER
SHIFTS_RANGE = f"'الإعدادات والقوائم'!$H$2:$H${1+len(SHIFTS)}"

# قائمة منسدلة لاختيار وردية بداية الفريق (B7) + تلوينها
dv_start = DataValidation(type="list", formula1=f"={SHIFTS_RANGE}", allow_blank=True)
ws_set.add_data_validation(dv_start)
dv_start.add("B7")
for nm in SHIFT_NAMES:
    ws_set.conditional_formatting.add(
        "B7", CellIsRule(operator="equal", formula=[f'"{nm}"'], fill=cf_fill(COLORS[nm])))

# جدول العطل الرسمية (قابل للتعديل/الإضافة)
HOL_FIRST, HOL_ROWS = 2, 20
HOL_LAST = HOL_FIRST + HOL_ROWS - 1
ws_set["K1"] = "تاريخ العطلة"
ws_set["L1"] = "المناسبة"
ws_set.column_dimensions["K"].width = 16
ws_set.column_dimensions["L"].width = 26
for cc in (ws_set["K1"], ws_set["L1"]):
    style_header(cc)
for i in range(HOL_ROWS):
    r = HOL_FIRST + i
    dcell = ws_set.cell(row=r, column=11)
    dcell.number_format = "DD/MM/YYYY"
    dcell.alignment = CENTER
    dcell.border = BORDER
    lcell = ws_set.cell(row=r, column=12)
    lcell.alignment = RIGHT
    lcell.border = BORDER
    lcell.font = BASE_FONT
    if i < len(HOLIDAYS):
        dcell.value = HOLIDAYS[i][0]
        lcell.value = HOLIDAYS[i][1]
        dcell.fill = PatternFill("solid", fgColor=COLORS["عطلة"])
HOLIDAYS_RANGE = f"'الإعدادات والقوائم'!$K${HOL_FIRST}:$K${HOL_LAST}"
wb.defined_names["Holidays"] = DefinedName("Holidays", attr_text=HOLIDAYS_RANGE)
HOLIDAYS_REF = "Holidays"

# =================================================================  الموظفون
ws_emp = wb.create_sheet("الموظفون")
ws_emp.sheet_view.rightToLeft = True
emp_headers = ["م", "الاسم", "الرقم الوظيفي", "بداية الدورة"]
widths = [6, 32, 16, 16]
for col, (h, w) in enumerate(zip(emp_headers, widths), start=1):
    c = ws_emp.cell(row=1, column=col, value=h)
    style_header(c)
    ws_emp.column_dimensions[get_column_letter(col)].width = w
for i, (num, name, jobno) in enumerate(EMPLOYEES, start=2):
    ws_emp.cell(row=i, column=1, value=num).alignment = CENTER
    ws_emp.cell(row=i, column=2, value=name).alignment = RIGHT
    ws_emp.cell(row=i, column=3, value=jobno).alignment = CENTER
    d = ws_emp.cell(row=i, column=4, value=PERIOD_START)
    d.alignment = CENTER
    d.number_format = "DD/MM/YYYY"
    for col in range(1, 5):
        ws_emp.cell(row=i, column=col).border = BORDER
        ws_emp.cell(row=i, column=col).font = BASE_FONT
ws_emp.freeze_panes = "A2"
EMP_LAST = 1 + len(EMPLOYEES)
wb.defined_names["EmpNames"] = DefinedName("EmpNames", attr_text=f"'الموظفون'!$B$2:$B${EMP_LAST}")
NAMES_RANGE = "EmpNames"

# =================================================================  حجز الإجازات
ws_lv = wb.create_sheet("حجز الإجازات")
ws_lv.sheet_view.rightToLeft = True
lv_headers = ["رقم الطلب", "اسم الموظف", "نوع الإجازة", "من تاريخ", "إلى تاريخ",
              "عدد الأيام", "الحالة", "فحص التعارض", "تنبيه التغطية", "ملاحظات"]
lv_widths = [10, 28, 14, 13, 13, 10, 14, 16, 18, 26]
for col, (h, w) in enumerate(zip(lv_headers, lv_widths), start=1):
    c = ws_lv.cell(row=4, column=col, value=h)
    style_header(c)
    ws_lv.column_dimensions[get_column_letter(col)].width = w

ws_lv["A1"] = "حجز الإجازات — قسم العمليات الجمركية"
ws_lv["A1"].font = TITLE_FONT
ws_lv["A2"] = "سجّل الطلب ثم تأكد من خانتي «فحص التعارض» و«تنبيه التغطية» قبل الاعتماد."
ws_lv["A2"].font = Font(name="Arial", italic=True, color="C00000", size=10)

# نطاقات مطلقة لصفوف الإجازات
B = lambda r: f"$B${r}"
EMPS = f"'حجز الإجازات'!$B${L_FIRST}:$B${L_LAST}"
TYPES_COL = f"'حجز الإجازات'!$C${L_FIRST}:$C${L_LAST}"
STARTS = f"'حجز الإجازات'!$D${L_FIRST}:$D${L_LAST}"
ENDS = f"'حجز الإجازات'!$E${L_FIRST}:$E${L_LAST}"
STAT = f"'حجز الإجازات'!$G${L_FIRST}:$G${L_LAST}"

for row in range(L_FIRST, L_LAST + 1):
    # رقم الطلب
    ws_lv.cell(row=row, column=1, value=row - L_FIRST + 1).alignment = CENTER
    # عدد الأيام = (نهاية - بداية + 1) مطروحاً منها العطل الرسمية ضمن الفترة
    ws_lv.cell(row=row, column=6,
               value=(f'=IF(OR($D{row}="",$E{row}=""),"",$E{row}-$D{row}+1'
                      f'-SUMPRODUCT(({HOLIDAYS_REF}>=$D{row})*({HOLIDAYS_REF}<=$E{row})))')
               ).alignment = CENTER
    # فحص التعارض: تداخل مع طلب آخر لنفس الموظف غير مرفوض
    overlap = (f'SUMPRODUCT(({EMPS}=$B{row})*({STAT}<>"مرفوض")'
               f'*({STARTS}<=$E{row})*($D{row}<=({ENDS})))')
    self_q = f'(($B{row}<>"")*($G{row}<>"مرفوض"))'
    ws_lv.cell(row=row, column=8,
               value=f'=IF(OR($B{row}="",$D{row}="",$E{row}=""),"",'
                     f'IF({overlap}-{self_q}>0,"⚠ تعارض","✓ سليم"))').alignment = CENTER
    # تنبيه التغطية: عدد المُجازين (غير مرفوض) المتداخلين مع يوم البداية
    cover = (f'SUMPRODUCT(({STAT}<>"مرفوض")*({STARTS}<=$D{row})*($D{row}<=({ENDS})))')
    ws_lv.cell(row=row, column=9,
               value=f'=IF($D{row}="","",IF({cover}>{MAX_LEAVE},'
                     f'"⚠ مجازين "&{cover}&" / الحد "&{MAX_LEAVE},'
                     f'"✓ مجازين "&{cover}&" / الحد "&{MAX_LEAVE}))').alignment = CENTER
    for col in range(1, 11):
        cell = ws_lv.cell(row=row, column=col)
        cell.border = BORDER
        cell.font = BASE_FONT
        if col in (4, 5):
            cell.number_format = "DD/MM/YYYY"
        if col in (2, 10):
            cell.alignment = RIGHT
        elif col not in (4, 5):
            cell.alignment = CENTER

# تعبئة بيانات الإجازات المسجّلة مسبقاً
for i, (name, typ, s, e, st) in enumerate(LEAVE_DATA):
    row = L_FIRST + i
    ws_lv.cell(row=row, column=2, value=name)
    ws_lv.cell(row=row, column=3, value=typ)
    ws_lv.cell(row=row, column=4, value=s).number_format = "DD/MM/YYYY"
    ws_lv.cell(row=row, column=5, value=e).number_format = "DD/MM/YYYY"
    ws_lv.cell(row=row, column=7, value=st)

# قوائم منسدلة
dv_name = DataValidation(type="list", formula1=f"={NAMES_RANGE}", allow_blank=True)
dv_type = DataValidation(type="list", formula1=f"={TYPES_RANGE}", allow_blank=True)
dv_stat = DataValidation(type="list", formula1=f"={STATUS_RANGE}", allow_blank=True)
ws_lv.add_data_validation(dv_name)
ws_lv.add_data_validation(dv_type)
ws_lv.add_data_validation(dv_stat)
dv_name.add(f"B{L_FIRST}:B{L_LAST}")
dv_type.add(f"C{L_FIRST}:C{L_LAST}")
dv_stat.add(f"G{L_FIRST}:G{L_LAST}")

# تنسيق شرطي لورقة الإجازات
ws_lv.conditional_formatting.add(
    f"H{L_FIRST}:H{L_LAST}",
    CellIsRule(operator="equal", formula=['"⚠ تعارض"'],
               fill=cf_fill("FFC7CE"),
               font=Font(color="9C0006", bold=True)))
ws_lv.conditional_formatting.add(
    f"H{L_FIRST}:H{L_LAST}",
    CellIsRule(operator="equal", formula=['"✓ سليم"'],
               fill=cf_fill("C6EFCE"),
               font=Font(color="006100")))
ws_lv.conditional_formatting.add(
    f"I{L_FIRST}:I{L_LAST}",
    FormulaRule(formula=[f'LEFT($I{L_FIRST})="⚠"'],
                fill=cf_fill("FFEB9C"),
                font=Font(color="9C6500", bold=True)))
# تلوين نوع الإجازة حسب اللون
for t in LEAVE_TYPES:
    ws_lv.conditional_formatting.add(
        f"C{L_FIRST}:C{L_LAST}",
        CellIsRule(operator="equal", formula=[f'"{t}"'],
                   fill=cf_fill(COLORS[t])))
ws_lv.freeze_panes = f"A{L_FIRST}"

# تحويل منطقة الطلبات إلى جدول إكسل ذكي (تنسيق احترافي + فلترة + امتداد تلقائي للصيغ)
lv_table = Table(displayName="جدول_الإجازات", ref=f"A4:J{L_LAST}")
lv_table.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
    showRowStripes=True, showColumnStripes=False)
ws_lv.add_table(lv_table)

# رابط رجوع إلى لوحة المعلومات
back = ws_lv.cell(row=2, column=8, value="↩ الرجوع للوحة المعلومات")
back.hyperlink = Hyperlink(ref="H2", location="'لوحة المعلومات'!A1",
                           display="الرجوع للوحة المعلومات")
back.font = Font(name="Arial", bold=True, color="2E75B6", underline="single", size=10)
back.alignment = CENTER

# =================================================================  جدول الورديات
ws_r = wb.create_sheet("جدول الورديات")
ws_r.sheet_view.rightToLeft = True
ws_r["A1"] = "جدول الورديات — 6 عمل / 4 راحة — صباح (6ص-1م) / عصر (1م-9م) / ليل (9م-6ص)"
ws_r["A1"].font = TITLE_FONT

# الأعمدة الثابتة
fixed = ["م", "الاسم", "الرقم الوظيفي"]
for col, h in enumerate(fixed, start=1):
    c = ws_r.cell(row=3, column=col, value=h)
    style_header(c)
ws_r.column_dimensions["A"].width = 5
ws_r.column_dimensions["B"].width = 26
ws_r.column_dimensions["C"].width = 13

DAY_COL0 = 4  # أول عمود تاريخ = D
dates = [PERIOD_START + datetime.timedelta(days=i) for i in range(DAYS)]
# صف اليوم (الاسم) في الصف 2، والتاريخ في الصف 3 يستخدم كعنوان؟ نضع التاريخ بصف 3 العلوي
# نستخدم: صف2 = اسم اليوم ، صف3 = التاريخ (وهو الذي تشير له الصيغ)
for i, d in enumerate(dates):
    col = DAY_COL0 + i
    letter = get_column_letter(col)
    ws_r.column_dimensions[letter].width = 6
    dn = ws_r.cell(row=2, column=col, value=AR_DAYS[d.weekday()])
    dn.fill = PatternFill("solid", fgColor="8EA9DB")
    dn.font = Font(name="Arial", bold=True, color="1F3864", size=8)
    dn.alignment = CENTER
    dn.border = BORDER
    dc = ws_r.cell(row=3, column=col, value=d)
    dc.number_format = "DD/MM"
    style_header(dc)
    if d.weekday() == 4:  # الجمعة
        dc.fill = PatternFill("solid", fgColor="C00000")
    if d.day == 1 or i == 0:
        ml = ws_r.cell(row=1, column=col, value=f"{AR_MONTHS[d.month-1]} {d.year}")
        ml.font = Font(name="Arial", bold=True, color="1F3864", size=10)
        ml.alignment = Alignment(horizontal="right", vertical="center")

ws_r.cell(row=2, column=2, value="التاريخ ←").alignment = CENTER

EMP_ROW0 = 4
for idx, (num, name, jobno) in enumerate(EMPLOYEES):
    row = EMP_ROW0 + idx
    erow = 2 + idx  # صف الموظف في ورقة الموظفين
    ws_r.cell(row=row, column=1, value=num).alignment = CENTER
    ws_r.cell(row=row, column=2, value=f"='الموظفون'!B{erow}").alignment = RIGHT
    ws_r.cell(row=row, column=3, value=f"='الموظفون'!C{erow}").alignment = CENTER
    cs = f"'الموظفون'!$D${erow}"          # بداية الدورة لهذا الموظف
    nm = f"'الموظفون'!$B${erow}"          # اسم الموظف
    for i in range(DAYS):
        col = DAY_COL0 + i
        dref = f"{get_column_letter(col)}$3"
        # داخل دورة العمل: يومان صباح + يومان عصر + يومان ليل (يتكرر النمط كل دورة)
        rot = (f'INDEX({SHIFTS_RANGE},MOD(INT(MOD({dref}-{cs},{CYCLE_WORK}+{CYCLE_REST})'
               f'*3/{CYCLE_WORK})+MATCH({START_SHIFT},{SHIFTS_RANGE},0)-1,3)+1)')
        # حالة الدورة: عطلة رسمية / وردية الفريق في أيام العمل / راحة في أيام الراحة
        cycle = (f'IF({dref}<{cs},"",IF(ISNUMBER(MATCH({dref},{HOLIDAYS_REF},0)),"عطلة",'
                 f'IF(MOD({dref}-{cs},{CYCLE_WORK}+{CYCLE_REST})<{CYCLE_WORK},{rot},"راحة")))')
        # هل توجد إجازة معتمدة لهذا الموظف بهذا التاريخ؟
        match = (f'SUMPRODUCT(({EMPS}={nm})*({STARTS}<={dref})*'
                 f'(({ENDS})>={dref})*({STAT}<>"مرفوض")*ROW({EMPS}))')
        idxf = f'({match})-{L_FIRST}+1'
        formula = (f'=IF({match}=0,{cycle},INDEX({TYPES_COL},{idxf}))')
        cc = ws_r.cell(row=row, column=col, value=formula)
        cc.alignment = CENTER
        cc.font = Font(name="Arial", size=8)
        cc.border = BORDER
    for col in (1, 2, 3):
        ws_r.cell(row=row, column=col).border = BORDER
        ws_r.cell(row=row, column=col).font = BASE_FONT

# صفوف الملخص أسفل الجدول (عدّاد لكل وردية + الإجمالي + المُجازين)
emp_first_letter = get_column_letter(DAY_COL0)
emp_last_row = EMP_ROW0 + len(EMPLOYEES) - 1
summary_labels = SHIFT_NAMES + ["إجمالي العاملين", "عدد المُجازين"]
summary_rows = {}
for k, lbl in enumerate(summary_labels):
    rr = emp_last_row + 1 + k
    summary_rows[lbl] = rr
    lc = ws_r.cell(row=rr, column=2, value=lbl)
    lc.alignment = RIGHT
    lc.font = Font(name="Arial", bold=True, size=10)
    lc.border = BORDER
    if lbl in COLORS:
        lc.fill = PatternFill("solid", fgColor=COLORS[lbl])

total_row = summary_rows["إجمالي العاملين"]
for i in range(DAYS):
    col = DAY_COL0 + i
    letter = get_column_letter(col)
    rng = f"{letter}{EMP_ROW0}:{letter}{emp_last_row}"
    # عدّاد كل وردية
    for nm in SHIFT_NAMES:
        rr = summary_rows[nm]
        cc = ws_r.cell(row=rr, column=col, value=f'=COUNTIF({rng},"{nm}")')
        cc.alignment = CENTER
        cc.font = Font(name="Arial", bold=True, size=9)
        cc.border = BORDER
    # إجمالي العاملين = مجموع الورديات الثلاث
    parts = "+".join(f'COUNTIF({rng},"{nm}")' for nm in SHIFT_NAMES)
    ct = ws_r.cell(row=total_row, column=col, value=f"={parts}")
    ct.alignment = CENTER
    ct.font = Font(name="Arial", bold=True, size=9)
    ct.border = BORDER
    ws_r.conditional_formatting.add(
        ct.coordinate,
        CellIsRule(operator="lessThan", formula=[MIN_STAFF],
                   fill=cf_fill("FFC7CE"),
                   font=Font(color="9C0006", bold=True)))
    # المُجازين = خلايا ليست وردية/راحة/فارغة
    excl = "".join(f'*({rng}<>"{nm}")' for nm in SHIFT_NAMES)
    cl = ws_r.cell(row=summary_rows["عدد المُجازين"], column=col,
                   value=f'=SUMPRODUCT(({rng}<>""){excl}*({rng}<>"راحة"))')
    cl.alignment = CENTER
    cl.font = Font(name="Arial", bold=True, size=9)
    cl.border = BORDER

sum_row2 = summary_rows["عدد المُجازين"]

# تنسيق شرطي لخلايا الجدول (ألوان الحالات)
grid_range = (f"{emp_first_letter}{EMP_ROW0}:"
              f"{get_column_letter(DAY_COL0+DAYS-1)}{EMP_ROW0+len(EMPLOYEES)-1}")
for key, color in COLORS.items():
    ws_r.conditional_formatting.add(
        grid_range,
        CellIsRule(operator="equal", formula=[f'"{key}"'],
                   fill=cf_fill(color)))

ws_r.freeze_panes = "D4"

# مفتاح الألوان (Legend)
leg_row = sum_row2 + 2
ws_r.cell(row=leg_row, column=2, value="مفتاح الألوان:").font = Font(bold=True, size=10)
legend_items = SHIFT_NAMES + ["راحة", "عطلة"] + LEAVE_TYPES
for i, k in enumerate(legend_items):
    cell = ws_r.cell(row=leg_row + 1 + i, column=2, value=k)
    cell.fill = PatternFill("solid", fgColor=COLORS[k])
    cell.alignment = RIGHT
    cell.border = BORDER
    cell.font = BASE_FONT

# =================================================================  تقويم الإجازات (عرض واضح للإجازات فقط)
ws_lc = wb.create_sheet("تقويم الإجازات")
ws_lc.sheet_view.rightToLeft = True
ws_lc["A1"] = "تقويم الإجازات — يعرض كل الإجازات (عدا المرفوضة) بألوانها"
ws_lc["A1"].font = TITLE_FONT
ws_lc["A2"] = "كل خلية ملوّنة = إجازة لذلك اليوم. الخلايا الفارغة = لا توجد إجازة."
ws_lc["A2"].font = Font(name="Arial", italic=True, color="44546A", size=10)

for col, h in enumerate(["م", "الاسم", "الرقم الوظيفي"], start=1):
    style_header(ws_lc.cell(row=3, column=col, value=h))
ws_lc.column_dimensions["A"].width = 5
ws_lc.column_dimensions["B"].width = 26
ws_lc.column_dimensions["C"].width = 13

for i, d in enumerate(dates):
    col = DAY_COL0 + i
    letter = get_column_letter(col)
    ws_lc.column_dimensions[letter].width = 6
    dn = ws_lc.cell(row=2, column=col, value=AR_DAYS[d.weekday()])
    dn.fill = PatternFill("solid", fgColor="8EA9DB")
    dn.font = Font(name="Arial", bold=True, color="1F3864", size=8)
    dn.alignment = CENTER
    dn.border = BORDER
    dc = ws_lc.cell(row=3, column=col, value=d)
    dc.number_format = "DD/MM"
    style_header(dc)
    if d.weekday() == 4:
        dc.fill = PatternFill("solid", fgColor="C00000")
    if d.day == 1 or i == 0:
        ml = ws_lc.cell(row=1, column=col, value=f"{AR_MONTHS[d.month-1]} {d.year}")
        ml.font = Font(name="Arial", bold=True, color="1F3864", size=10)
        ml.alignment = Alignment(horizontal="right", vertical="center")

TOTAL_COL = DAY_COL0 + DAYS            # عمود إجمالي أيام الإجازة لكل موظف
tcl = ws_lc.cell(row=3, column=TOTAL_COL, value="إجمالي الأيام")
style_header(tcl)
ws_lc.column_dimensions[get_column_letter(TOTAL_COL)].width = 11

for idx, (num, name, jobno) in enumerate(EMPLOYEES):
    row = EMP_ROW0 + idx
    erow = 2 + idx
    ws_lc.cell(row=row, column=1, value=num).alignment = CENTER
    ws_lc.cell(row=row, column=2, value=f"='الموظفون'!B{erow}").alignment = RIGHT
    ws_lc.cell(row=row, column=3, value=f"='الموظفون'!C{erow}").alignment = CENTER
    nm = f"'الموظفون'!$B${erow}"
    for i in range(DAYS):
        col = DAY_COL0 + i
        dref = f"{get_column_letter(col)}$3"
        match = (f'SUMPRODUCT(({EMPS}={nm})*({STARTS}<={dref})*'
                 f'(({ENDS})>={dref})*({STAT}<>"مرفوض")*ROW({EMPS}))')
        idxf = f'({match})-{L_FIRST}+1'
        cc = ws_lc.cell(row=row, column=col,
                        value=f'=IF({match}=0,"",INDEX({TYPES_COL},{idxf}))')
        cc.alignment = CENTER
        cc.font = Font(name="Arial", size=8)
        cc.border = BORDER
    # إجمالي أيام الإجازة لهذا الموظف خلال الفترة
    rrow = (f"{get_column_letter(DAY_COL0)}{row}:"
            f"{get_column_letter(DAY_COL0+DAYS-1)}{row}")
    tc = ws_lc.cell(row=row, column=TOTAL_COL, value=f'=SUMPRODUCT(--({rrow}<>""))')
    tc.alignment = CENTER
    tc.font = Font(name="Arial", bold=True, size=10)
    tc.border = BORDER
    for col in (1, 2, 3):
        ws_lc.cell(row=row, column=col).border = BORDER
        ws_lc.cell(row=row, column=col).font = BASE_FONT

# صف إجمالي المُجازين لكل يوم
lc_total_row = EMP_ROW0 + len(EMPLOYEES)
tl = ws_lc.cell(row=lc_total_row, column=2, value="عدد المُجازين")
tl.alignment = RIGHT
tl.font = Font(name="Arial", bold=True, size=10)
tl.border = BORDER
for i in range(DAYS):
    col = DAY_COL0 + i
    letter = get_column_letter(col)
    colrng = f"{letter}{EMP_ROW0}:{letter}{EMP_ROW0+len(EMPLOYEES)-1}"
    tc = ws_lc.cell(row=lc_total_row, column=col, value=f'=SUMPRODUCT(--({colrng}<>""))')
    tc.alignment = CENTER
    tc.font = Font(name="Arial", bold=True, size=9)
    tc.border = BORDER
    # تلوين الأيام التي يتجاوز فيها عدد المُجازين الحد المسموح
    ws_lc.conditional_formatting.add(
        tc.coordinate,
        CellIsRule(operator="greaterThan", formula=[MAX_LEAVE],
                   fill=cf_fill("FFC7CE"),
                   font=Font(color="9C0006", bold=True)))

# تلوين خلايا الإجازات حسب النوع
lc_grid = (f"{get_column_letter(DAY_COL0)}{EMP_ROW0}:"
           f"{get_column_letter(DAY_COL0+DAYS-1)}{EMP_ROW0+len(EMPLOYEES)-1}")
for t in LEAVE_TYPES:
    ws_lc.conditional_formatting.add(
        lc_grid,
        CellIsRule(operator="equal", formula=[f'"{t}"'],
                   fill=cf_fill(COLORS[t])))
ws_lc.freeze_panes = "D4"

# مفتاح ألوان أنواع الإجازات
lc_leg = lc_total_row + 2
ws_lc.cell(row=lc_leg, column=2, value="مفتاح ألوان الإجازات:").font = Font(bold=True, size=10)
for i, t in enumerate(LEAVE_TYPES):
    cell = ws_lc.cell(row=lc_leg + 1 + i, column=2, value=t)
    cell.fill = PatternFill("solid", fgColor=COLORS[t])
    cell.alignment = RIGHT
    cell.border = BORDER
    cell.font = BASE_FONT

# =================================================================  لوحة المعلومات (داشبورد)
ws_db = wb.create_sheet("لوحة المعلومات")
ws_db.sheet_view.rightToLeft = True
ws_db.sheet_view.showGridLines = False
for c in range(1, 18):
    ws_db.column_dimensions[get_column_letter(c)].width = 9
ws_db.column_dimensions["A"].width = 2
for g in ("E", "I", "M"):
    ws_db.column_dimensions[g].width = 2

# مراجع ورقة الإجازات
LV_B = "'حجز الإجازات'!$B$5:$B$104"
LV_C = "'حجز الإجازات'!$C$5:$C$104"
LV_D = "'حجز الإجازات'!$D$5:$D$104"
LV_E = "'حجز الإجازات'!$E$5:$E$104"
LV_F = "'حجز الإجازات'!$F$5:$F$104"
LV_G = "'حجز الإجازات'!$G$5:$G$104"
LV_H = "'حجز الإجازات'!$H$5:$H$104"
_dic = "MOD(TODAY()-'الإعدادات والقوائم'!$B$2," + f"{CYCLE_WORK}+{CYCLE_REST})"
TEAM_TODAY = (f'IF(TODAY()<\'الإعدادات والقوائم\'!$B$2,"-",IF({_dic}>={CYCLE_WORK},"راحة",'
              f'INDEX({SHIFTS_RANGE},MOD(INT({_dic}*3/{CYCLE_WORK})'
              f'+MATCH({START_SHIFT},{SHIFTS_RANGE},0)-1,3)+1)))')

# --- العنوان ---
ws_db.merge_cells("B1:P2")
t = ws_db["B1"]
t.value = "لوحة المعلومات — جدول الورديات وحجز الإجازات"
t.font = Font(name="Arial", bold=True, size=20, color="FFFFFF")
t.alignment = CENTER
t.fill = PatternFill("solid", fgColor="1F3864")
ws_db.merge_cells("B3:P3")
s = ws_db["B3"]
s.value = '="قسم العمليات الجمركية     •     آخر تحديث: "&TEXT(TODAY(),"DD/MM/YYYY")'
s.font = Font(name="Arial", bold=True, size=11, color="44546A")
s.alignment = CENTER


def card(col0, row, label, formula, color, valcolor="1F3864"):
    c0 = get_column_letter(col0)
    c1 = get_column_letter(col0 + 2)
    ws_db.merge_cells(f"{c0}{row}:{c1}{row}")
    lc = ws_db[f"{c0}{row}"]
    lc.value = label
    lc.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    lc.alignment = CENTER
    lc.fill = PatternFill("solid", fgColor="44546A")
    lc.border = BORDER
    ws_db.merge_cells(f"{c0}{row+1}:{c1}{row+1}")
    vc = ws_db[f"{c0}{row+1}"]
    vc.value = formula
    vc.font = Font(name="Arial", bold=True, size=24, color=valcolor)
    vc.alignment = CENTER
    vc.fill = PatternFill("solid", fgColor=color)
    vc.border = BORDER
    ws_db.row_dimensions[row + 1].height = 34


cards1 = [
    (2,  "عدد الموظفين", f"=COUNTA({NAMES_RANGE})", "DDEBF7"),
    (6,  "إجمالي الطلبات", f"=COUNTA({LV_B})", "DDEBF7"),
    (10, "الطلبات المعتمدة", f'=COUNTIF({LV_G},"معتمد")', "E2EFDA"),
    (14, "طلبات فيها تعارض", f'=COUNTIF({LV_H},"*تعارض*")', "FCE4D6"),
]
for col0, lbl, frm, clr in cards1:
    card(col0, 5, lbl, frm, clr)
cards2 = [
    (2,  "قيد الانتظار", f'=COUNTIF({LV_G},"قيد الانتظار")', "FFF2CC"),
    (6,  "الطلبات المرفوضة", f'=COUNTIF({LV_G},"مرفوض")', "F2F2F2"),
    (10, "إجمالي أيام الإجازات", f'=SUMIFS({LV_F},{LV_G},"معتمد")', "E2EFDA"),
    (14, "مُجازون اليوم",
     f'=SUMPRODUCT(({LV_G}="معتمد")*({LV_D}<=TODAY())*({LV_E}>=TODAY()))', "DDEBF7"),
]
for col0, lbl, frm, clr in cards2:
    card(col0, 8, lbl, frm, clr)
# تمييز بطاقة التعارض بالأحمر عند وجود تعارض
ws_db.conditional_formatting.add(
    "N6:P6",
    CellIsRule(operator="greaterThan", formula=["0"], fill=cf_fill("FFC7CE"),
               font=Font(bold=True, size=24, color="9C0006")))

# --- وردية اليوم + زر تسجيل إجازة ---
ws_db.merge_cells("B11:G11")
lc = ws_db["B11"]
lc.value = "وردية الفريق اليوم"
lc.font = Font(bold=True, size=12, color="FFFFFF")
lc.alignment = CENTER
lc.fill = PatternFill("solid", fgColor="2E75B6")
lc.border = BORDER
ws_db.merge_cells("B12:G12")
vc = ws_db["B12"]
vc.value = f"={TEAM_TODAY}"
vc.font = Font(bold=True, size=22, color="1F3864")
vc.alignment = CENTER
vc.fill = PatternFill("solid", fgColor="DDEBF7")
vc.border = BORDER
ws_db.row_dimensions[12].height = 36

ws_db.merge_cells("J11:P12")
btn = ws_db["J11"]
btn.value = "➕  تسجيل إجازة جديدة  (اضغط هنا)"
btn.hyperlink = Hyperlink(ref="J11", location="'حجز الإجازات'!B5",
                          display="تسجيل إجازة جديدة")
btn.font = Font(bold=True, size=15, color="FFFFFF", underline="single")
btn.alignment = CENTER
btn.fill = PatternFill("solid", fgColor="548235")
for rr in (11, 12):
    for cc in range(10, 17):
        ws_db.cell(row=rr, column=cc).border = BORDER


def tbl_header(row, cols_labels):
    for col, h in cols_labels:
        cc = ws_db.cell(row=row, column=col, value=h)
        cc.fill = HDR_FILL
        cc.font = HDR_FONT
        cc.alignment = CENTER
        cc.border = BORDER


def data_cell(row, col, val, align=CENTER, fill=None):
    cc = ws_db.cell(row=row, column=col, value=val)
    cc.alignment = align
    cc.border = BORDER
    cc.font = BASE_FONT
    if fill:
        cc.fill = PatternFill("solid", fgColor=fill)
    return cc


# --- الإجازات حسب النوع ---
ws_db.cell(row=14, column=2, value="📊 الإجازات حسب النوع").font = Font(bold=True, size=12, color="1F3864")
tbl_header(15, [(2, "النوع"), (3, "عدد الطلبات"), (4, "إجمالي الأيام")])
for i, ty in enumerate(LEAVE_TYPES):
    r = 16 + i
    data_cell(r, 2, ty, align=RIGHT, fill=COLORS[ty])
    data_cell(r, 3, f'=COUNTIF({LV_C},"{ty}")')
    data_cell(r, 4, f'=SUMIF({LV_C},"{ty}",{LV_F})')
ch1 = BarChart()
ch1.type = "col"
ch1.title = "أيام الإجازة حسب النوع"
ch1.height = 7
ch1.width = 12
ch1.add_data(Reference(ws_db, min_col=4, min_row=15, max_row=15 + len(LEAVE_TYPES)), titles_from_data=True)
ch1.set_categories(Reference(ws_db, min_col=2, min_row=16, max_row=15 + len(LEAVE_TYPES)))
ch1.legend = None
ws_db.add_chart(ch1, "F15")

# --- إجمالي أيام الإجازة لكل موظف ---
ws_db.cell(row=31, column=2, value="📊 إجمالي أيام الإجازة لكل موظف").font = Font(bold=True, size=12, color="1F3864")
tbl_header(32, [(2, "الموظف"), (3, "عدد الطلبات"), (4, "إجمالي الأيام")])
for i in range(len(EMPLOYEES)):
    r = 33 + i
    erow = 2 + i
    data_cell(r, 2, f"='الموظفون'!B{erow}", align=RIGHT)
    data_cell(r, 3, f'=COUNTIFS({LV_B},\'الموظفون\'!B{erow},{LV_G},"<>مرفوض")')
    data_cell(r, 4, f'=SUMIFS({LV_F},{LV_B},\'الموظفون\'!B{erow},{LV_G},"<>مرفوض")')
ch2 = BarChart()
ch2.type = "bar"
ch2.title = "إجمالي أيام الإجازة لكل موظف"
ch2.height = 7.5
ch2.width = 12
ch2.add_data(Reference(ws_db, min_col=4, min_row=32, max_row=32 + len(EMPLOYEES)), titles_from_data=True)
ch2.set_categories(Reference(ws_db, min_col=2, min_row=33, max_row=32 + len(EMPLOYEES)))
ch2.legend = None
ws_db.add_chart(ch2, "F31")

# --- توزيع الطلبات حسب الحالة ---
ws_db.cell(row=47, column=2, value="🥧 توزيع الطلبات حسب الحالة").font = Font(bold=True, size=12, color="1F3864")
status_colors = {"معتمد": "C6EFCE", "قيد الانتظار": "FFEB9C", "مرفوض": "FFC7CE"}
tbl_header(48, [(2, "الحالة"), (3, "عدد الطلبات")])
for i, stt in enumerate(STATUSES):
    r = 49 + i
    data_cell(r, 2, stt, align=RIGHT, fill=status_colors[stt])
    data_cell(r, 3, f'=COUNTIF({LV_G},"{stt}")')
ch3 = PieChart()
ch3.title = "توزيع الطلبات حسب الحالة"
ch3.height = 7
ch3.width = 10
ch3.add_data(Reference(ws_db, min_col=3, min_row=48, max_row=48 + len(STATUSES)), titles_from_data=True)
ch3.set_categories(Reference(ws_db, min_col=2, min_row=49, max_row=48 + len(STATUSES)))
ws_db.add_chart(ch3, "F47")
# =================================================================  كشف يومي للطباعة
ws_day = wb.create_sheet("كشف يومي")
ws_day.sheet_view.rightToLeft = True
ws_day.sheet_view.showGridLines = False
for c, w in {"A": 5, "B": 28, "C": 14, "D": 16, "E": 18, "F": 18}.items():
    ws_day.column_dimensions[c].width = w
ws_day.merge_cells("A1:F1")
h = ws_day["A1"]
h.value = "كشف الوردية اليومي — قسم العمليات الجمركية"
h.font = Font(bold=True, size=16, color="FFFFFF")
h.alignment = CENTER
h.fill = PatternFill("solid", fgColor="1F3864")
ws_day.row_dimensions[1].height = 24
ws_day["B2"] = "📅 التاريخ:"
ws_day["B2"].font = Font(bold=True, size=12)
ws_day["B2"].alignment = RIGHT
dt = ws_day["C2"]
dt.value = "=TODAY()"
dt.number_format = "DD/MM/YYYY"
dt.font = Font(bold=True, size=12, color="1F3864")
dt.alignment = CENTER
dt.fill = PatternFill("solid", fgColor="FFF2CC")
dt.border = BORDER
ws_day["B3"] = "اليوم:"
ws_day["B3"].font = Font(bold=True)
ws_day["B3"].alignment = RIGHT
wk = ws_day["C3"]
wk.value = ('=CHOOSE(WEEKDAY($C$2,2),"الإثنين","الثلاثاء","الأربعاء","الخميس",'
            '"الجمعة","السبت","الأحد")')
wk.alignment = CENTER
wk.font = Font(bold=True)
ws_day["B4"] = "💡 غيّر التاريخ في الخانة الصفراء ليتحدّث الكشف، ثم اطبع."
ws_day["B4"].font = Font(italic=True, size=9, color="44546A")
ws_day["B4"].alignment = RIGHT
DT = "$C$2"
hdr = ["م", "الاسم", "الرقم الوظيفي", "الوردية/الحالة", "توقيع الحضور", "توقيع الانصراف"]
for c, ttl in enumerate(hdr, start=1):
    style_header(ws_day.cell(row=5, column=c, value=ttl))
for i, (num, name, job) in enumerate(EMPLOYEES):
    r = 6 + i
    erow = 2 + i
    ws_day.cell(row=r, column=1, value=num).alignment = CENTER
    ws_day.cell(row=r, column=2, value=f"='الموظفون'!B{erow}").alignment = RIGHT
    ws_day.cell(row=r, column=3, value=f"='الموظفون'!C{erow}").alignment = CENTER
    cs = f"'الموظفون'!$D${erow}"
    nm = f"'الموظفون'!$B${erow}"
    rot = (f'INDEX({SHIFTS_RANGE},MOD(INT(MOD({DT}-{cs},{CYCLE_WORK}+{CYCLE_REST})'
           f'*3/{CYCLE_WORK})+MATCH({START_SHIFT},{SHIFTS_RANGE},0)-1,3)+1)')
    cyc = (f'IF({DT}<{cs},"",IF(ISNUMBER(MATCH({DT},{HOLIDAYS_REF},0)),"عطلة",'
           f'IF(MOD({DT}-{cs},{CYCLE_WORK}+{CYCLE_REST})<{CYCLE_WORK},{rot},"راحة")))')
    match = (f'SUMPRODUCT(({EMPS}={nm})*({STARTS}<={DT})*(({ENDS})>={DT})'
             f'*({STAT}<>"مرفوض")*ROW({EMPS}))')
    sc = ws_day.cell(row=r, column=4,
                     value=f'=IF({match}=0,{cyc},INDEX({TYPES_COL},({match})-{L_FIRST}+1))')
    sc.alignment = CENTER
    sc.font = Font(bold=True, size=10)
    for c in range(1, 7):
        ws_day.cell(row=r, column=c).border = BORDER
        if ws_day.cell(row=r, column=c).font.size is None:
            ws_day.cell(row=r, column=c).font = BASE_FONT
day_grid = f"D6:D{5+len(EMPLOYEES)}"
for k, clr in COLORS.items():
    ws_day.conditional_formatting.add(
        day_grid, CellIsRule(operator="equal", formula=[f'"{k}"'], fill=cf_fill(clr)))
sig_row = 6 + len(EMPLOYEES) + 1
ws_day.cell(row=sig_row, column=2, value="توقيع المشرف:").font = Font(bold=True)
ws_day.cell(row=sig_row, column=2).alignment = RIGHT
ws_day.print_area = f"A1:F{sig_row+1}"
ws_day.page_setup.orientation = "portrait"
ws_day.page_setup.fitToWidth = 1
ws_day.sheet_properties.pageSetUpPr.fitToPage = True

# =================================================================  متابعة الإجازات
ws_mon = wb.create_sheet("متابعة الإجازات")
ws_mon.sheet_view.rightToLeft = True
ws_mon.sheet_view.showGridLines = False
for c, w in {"A": 5, "B": 28, "C": 18, "D": 18, "E": 14}.items():
    ws_mon.column_dimensions[c].width = w
ws_mon.merge_cells("A1:E1")
hm = ws_mon["A1"]
hm.value = "متابعة الإجازات — اليوم والقادمة"
hm.font = Font(bold=True, size=16, color="FFFFFF")
hm.alignment = CENTER
hm.fill = PatternFill("solid", fgColor="1F3864")
ws_mon.row_dimensions[1].height = 24
LV_E = "'حجز الإجازات'!$E$5:$E$104"
ws_mon["B2"] = "مُجازون اليوم:"
ws_mon["B2"].font = Font(bold=True)
ws_mon["B2"].alignment = RIGHT
ws_mon["C2"] = f'=SUMPRODUCT(({LV_G}="معتمد")*({LV_D}<=TODAY())*({LV_E}>=TODAY()))'
ws_mon["C2"].alignment = CENTER
ws_mon["C2"].font = Font(bold=True, size=12, color="1F3864")
ws_mon["B3"] = "إجازات تبدأ خلال 7 أيام:"
ws_mon["B3"].font = Font(bold=True)
ws_mon["B3"].alignment = RIGHT
ws_mon["C3"] = f'=SUMPRODUCT(({LV_G}<>"مرفوض")*({LV_D}>TODAY())*({LV_D}<=TODAY()+7))'
ws_mon["C3"].alignment = CENTER
ws_mon["C3"].font = Font(bold=True, size=12, color="1F3864")
mon_hdr = ["م", "الاسم", "حالة اليوم", "أقرب إجازة قادمة", "بعد (يوم)"]
for c, ttl in enumerate(mon_hdr, start=1):
    style_header(ws_mon.cell(row=5, column=c, value=ttl))
for i, (num, name, job) in enumerate(EMPLOYEES):
    r = 6 + i
    erow = 2 + i
    ws_mon.cell(row=r, column=1, value=num).alignment = CENTER
    ws_mon.cell(row=r, column=2, value=f"='الموظفون'!B{erow}").alignment = RIGHT
    cs = f"'الموظفون'!$D${erow}"
    nm = f"'الموظفون'!$B${erow}"
    rot = (f'INDEX({SHIFTS_RANGE},MOD(INT(MOD(TODAY()-{cs},{CYCLE_WORK}+{CYCLE_REST})'
           f'*3/{CYCLE_WORK})+MATCH({START_SHIFT},{SHIFTS_RANGE},0)-1,3)+1)')
    cyc = (f'IF(TODAY()<{cs},"",IF(ISNUMBER(MATCH(TODAY(),{HOLIDAYS_REF},0)),"عطلة",'
           f'IF(MOD(TODAY()-{cs},{CYCLE_WORK}+{CYCLE_REST})<{CYCLE_WORK},{rot},"راحة")))')
    match = (f'SUMPRODUCT(({EMPS}={nm})*({STARTS}<=TODAY())*(({ENDS})>=TODAY())'
             f'*({STAT}<>"مرفوض")*ROW({EMPS}))')
    sc = ws_mon.cell(row=r, column=3,
                     value=f'=IF({match}=0,{cyc},INDEX({TYPES_COL},({match})-{L_FIRST}+1))')
    sc.alignment = CENTER
    sc.font = Font(bold=True, size=10)
    nx = ws_mon.cell(row=r, column=4,
                     value=(f'=IFERROR(AGGREGATE(15,6,{LV_D}/(({LV_B}={nm})'
                            f'*({LV_D}>TODAY())*({LV_G}<>"مرفوض")),1),"")'))
    nx.number_format = "DD/MM/YYYY"
    nx.alignment = CENTER
    nx.font = BASE_FONT
    af = ws_mon.cell(row=r, column=5, value=f'=IF($D{r}="","",$D{r}-TODAY())')
    af.alignment = CENTER
    af.font = BASE_FONT
    for c in range(1, 6):
        ws_mon.cell(row=r, column=c).border = BORDER
mon_grid = f"C6:C{5+len(EMPLOYEES)}"
for k, clr in COLORS.items():
    ws_mon.conditional_formatting.add(
        mon_grid, CellIsRule(operator="equal", formula=[f'"{k}"'], fill=cf_fill(clr)))

# =================================================================  التوزيع العادل للإجازات
ws_fair = wb.create_sheet("التوزيع العادل")
ws_fair.sheet_view.rightToLeft = True
ws_fair.sheet_view.showGridLines = False
for c, w in {"A": 2, "B": 5, "C": 28, "D": 13, "E": 16, "F": 9, "G": 16, "H": 20}.items():
    ws_fair.column_dimensions[c].width = w
ws_fair.merge_cells("B1:H1")
hf = ws_fair["B1"]
hf.value = "التوزيع العادل للإجازات بين الفريق"
hf.font = Font(bold=True, size=16, color="FFFFFF")
hf.alignment = CENTER
hf.fill = PatternFill("solid", fgColor="1F3864")
ws_fair.row_dimensions[1].height = 24

FN, FL = 8, 8 + len(EMPLOYEES) - 1          # صفوف البيانات
DAYS_RNG = f"$E${FN}:$E${FL}"
# ملخص
summ = [
    ("متوسط أيام الإجازة للفريق", f"=ROUND(AVERAGE({DAYS_RNG}),1)"),
    ("الأكثر أخذاً", f"=INDEX($C${FN}:$C${FL},MATCH(MAX({DAYS_RNG}),{DAYS_RNG},0))"),
    ("الأقل أخذاً", f"=INDEX($C${FN}:$C${FL},MATCH(MIN({DAYS_RNG}),{DAYS_RNG},0))"),
]
for i, (lbl, frm) in enumerate(summ):
    r = 3 + i
    lc = ws_fair.cell(row=r, column=2, value=lbl)
    lc.value = lbl
    ws_fair.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    lc.font = Font(bold=True, size=11)
    lc.alignment = RIGHT
    lc.border = BORDER
    vc = ws_fair.cell(row=r, column=4, value=frm)
    ws_fair.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
    vc.font = Font(bold=True, size=12, color="1F3864")
    vc.alignment = CENTER
    vc.border = BORDER
    vc.fill = PatternFill("solid", fgColor="DDEBF7")

# رأس الجدول
fair_hdr = ["م", "الموظف", "عدد الطلبات", "إجمالي أيام الإجازة", "الترتيب",
            "الفرق عن المتوسط", "التقييم"]
for c, ttl in enumerate(fair_hdr, start=2):
    style_header(ws_fair.cell(row=7, column=c, value=ttl))
AVG_CELL = "$D$3"
for i in range(len(EMPLOYEES)):
    r = FN + i
    erow = 2 + i
    nm = f"'الموظفون'!$B${erow}"
    ws_fair.cell(row=r, column=2, value=i + 1).alignment = CENTER
    ws_fair.cell(row=r, column=3, value=f"='الموظفون'!B{erow}").alignment = RIGHT
    ws_fair.cell(row=r, column=4,
                 value=f'=COUNTIFS({LV_B},{nm},{LV_G},"معتمد")').alignment = CENTER
    ws_fair.cell(row=r, column=5,
                 value=f'=SUMIFS({LV_F},{LV_B},{nm},{LV_G},"معتمد")').alignment = CENTER
    ws_fair.cell(row=r, column=6,
                 value=f'=RANK($E{r},{DAYS_RNG},0)').alignment = CENTER
    ws_fair.cell(row=r, column=7,
                 value=f'=$E{r}-{AVG_CELL}').alignment = CENTER
    ws_fair.cell(row=r, column=8,
                 value=(f'=IF($E{r}>{AVG_CELL},"أعلى من المتوسط ⬆",'
                        f'IF($E{r}<{AVG_CELL},"أقل من المتوسط ⬇","ضمن المتوسط"))')).alignment = CENTER
    for c in range(2, 9):
        cell = ws_fair.cell(row=r, column=c)
        cell.border = BORDER
        if cell.font.size is None:
            cell.font = BASE_FONT

# تدرّج لوني على عمود الأيام: أخضر (أقل) ← أحمر (أكثر أخذاً)
from openpyxl.formatting.rule import ColorScaleRule
ws_fair.conditional_formatting.add(
    f"E{FN}:E{FL}",
    ColorScaleRule(start_type="min", start_color="C6EFCE",
                   mid_type="percentile", mid_value=50, mid_color="FFEB9C",
                   end_type="max", end_color="F8696B"))
# تمييز التقييم
ws_fair.conditional_formatting.add(
    f"H{FN}:H{FL}",
    FormulaRule(formula=[f'LEFT($H{FN},5)="أعلى"'], fill=cf_fill("FCE4D6")))
ws_fair.conditional_formatting.add(
    f"H{FN}:H{FL}",
    FormulaRule(formula=[f'LEFT($H{FN},4)="أقل"'], fill=cf_fill("E2EFDA")))

# رسم بياني
fair_chart = BarChart()
fair_chart.type = "bar"
fair_chart.title = "إجمالي أيام الإجازة لكل موظف (للعدالة)"
fair_chart.height = 8
fair_chart.width = 14
fair_chart.add_data(Reference(ws_fair, min_col=5, min_row=7, max_row=FL), titles_from_data=True)
fair_chart.set_categories(Reference(ws_fair, min_col=3, min_row=FN, max_row=FL))
fair_chart.legend = None
ws_fair.add_chart(fair_chart, "B18")

# =================================================================  (أُلغيت حماية الأوراق بناءً على طلب المستخدم — كل الخلايا قابلة للتعديل)

# ترتيب الداشبورد أولاً
wb.move_sheet("لوحة المعلومات", -wb.sheetnames.index("لوحة المعلومات"))

# ------------------------------------------------------------------ حفظ
# إجبار البرنامج على إعادة احتساب جميع الصيغ عند فتح الملف
try:
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.calcMode = "auto"
except Exception:
    pass
out = "/home/user/ghost/جدول_الورديات_وحجز_الإجازات.xlsx"
wb.save(out)
print("تم الحفظ:", out)
